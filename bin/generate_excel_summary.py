#!/usr/bin/env python3
import argparse
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from scipy.cluster.hierarchy import leaves_list, linkage


def main():
    args = parse_arguments()

    # load data
    df_samples = pd.read_csv(args.sample_summary).assign(
        sample=lambda x: x["sample"].astype(str),
        meandepth=lambda x: x.meandepth.round(),
        perc_covered=lambda x: x.perc_covered.round(1),
    )
    df_variants = get_variant_summary(args.variants_list)

    # create workbook
    wb = Workbook()

    ws_samples = wb.active
    ws_samples.title = "Samples"
    write_samples_sheet(ws_samples, df_samples, args.min_coverage)

    ws_variants = wb.create_sheet("Variants")
    write_variants_sheet(ws_variants, df_variants)

    wb.save(args.output)


def write_samples_sheet(sheet: Worksheet, df_samples: pd.DataFrame, min_coverage: int):
    """
    Write and format samples sheet in workbook
    """
    # write data
    for r in dataframe_to_rows(df_samples, header=True, index=False):
        sheet.append(r)

    # apply conditional filling to perc_covered column
    perc_column = next(
        x[0].column_letter for x in sheet.columns if x[0].value == "perc_covered"
    )
    for i, row in enumerate(df_samples.itertuples(), start=2):
        color = "cefbc1" if row.perc_covered > min_coverage else "fbc1c1"
        sheet[f"{perc_column}{i}"].fill = PatternFill(
            start_color=color, fill_type="solid"
        )

    # improve style
    for col in sheet.columns:
        col[0].font = Font(name="Calibri", bold=True)
        col[0].border = Border(bottom=Side(border_style="medium", color="000000"))
        col[0].alignment = Alignment(horizontal="center")
        sheet.column_dimensions[col[0].column_letter].width = 13


def write_variants_sheet(sheet: Worksheet, df_variants: pd.DataFrame):
    """
    Write and format variants sheet in workbook
    """
    # write data
    for r in dataframe_to_rows(df_variants, header=True, index=False):
        sheet.append(r)
    sheet.insert_rows(1)

    # apply conditional filling depending of variant presence
    for row in sheet.iter_rows(
        min_col=8,
        max_col=sheet.max_column,
        min_row=3,
        max_row=sheet.max_row,
    ):
        for cell in row:
            if cell.value == 1:
                cell.fill = PatternFill(start_color="3bbf97", fill_type="solid")
            cell.value = ""

    # improve style (common columns)
    column_widths = [8, 8, 8, 18, 10, 14, 14]
    for col, w in zip(sheet.iter_cols(max_row=2, max_col=7), column_widths):
        colname = col[0].column_letter
        col[0].value = col[1].value
        col[0].font = Font(name="Calibri", bold=True)
        col[0].border = Border(bottom=Side(border_style="medium", color="000000"))
        col[0].alignment = Alignment(horizontal="center")
        sheet.column_dimensions[colname].width = w
        sheet.merge_cells(f"{colname}1:{colname}2")

    # improve style (samples columns)
    sheet.row_dimensions[2].height = 60
    sheet["H1"].value = "samples"
    sheet["H1"].font = Font(name="Calibri", bold=True)
    sheet["H1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(
        start_row=1, end_row=1, start_column=8, end_column=sheet.max_column
    )
    for col in sheet.iter_cols(min_row=2, max_row=2, min_col=8):
        col[0].font = Font(name="Calibri", bold=True)
        col[0].border = Border(bottom=Side(border_style="medium", color="000000"))
        col[0].alignment = Alignment(horizontal="center", text_rotation=90)
        sheet.column_dimensions[col[0].column_letter].width = 3


def get_variant_summary(variant_list_filename: str) -> pd.DataFrame:
    """
    Load variant list and generate a summary with each variants as rows and samples as column

    :param variant_list_filename: filename of CSV file containing variant list
    :return: dataframe with the summary generated
    """
    variants_base_cols = [
        "pos",
        "ref",
        "alt",
        "variant_type",
        "gene",
        "nt_change",
        "aa_change",
    ]
    df_variants = pd.read_csv(variant_list_filename)
    variant_summary = (
        df_variants.assign(z=1)
        .pivot_table(index=["nt_change"], columns="sample", values="z", fill_value=0)
        .reset_index()
        .merge(df_variants[variants_base_cols].drop_duplicates())
        .sort_values(["pos", "alt"])
        .assign(
            nt_change=lambda x: x.nt_change.str.replace("[cn].", "", regex=True),
            aa_change=lambda x: x.aa_change.apply(
                lambda y: format_aminoacid_short(str(y))
            ),
        )
    )
    # leave gene field of intergenic variants empty
    variant_summary.loc[
        variant_summary.variant_type == "intergenic_region", "gene"
    ] = ""

    # sort columns
    variant_summary = variant_summary[
        variants_base_cols
        + get_sample_optimal_order(
            variant_summary[
                [x for x in variant_summary.columns if x not in variants_base_cols]
            ]
        )
    ]

    return variant_summary


def get_sample_optimal_order(sample_variants: pd.DataFrame) -> List[str]:
    """
    Use a hierarchical clustering to reorder samples

    :param sample_variants: datafram where each column is a sample
    :return list of sorted sample names (from the column names)
    """
    values = sample_variants.transpose()
    linked = linkage(values, method="single", optimal_ordering=True)
    names_sorted = [sample_variants.columns[x] for x in leaves_list(linked)]

    return names_sorted


def format_aminoacid_short(aa_change: str) -> str:
    """
    Format aminoacid change using short notation
    """
    if aa_change == "nan":
        return ""

    aa_list = {
        "Ala": "A",
        "Cys": "C",
        "Asp": "D",
        "Glu": "E",
        "Phe": "F",
        "Gly": "G",
        "His": "H",
        "Ile": "I",
        "Lys": "K",
        "Leu": "L",
        "Met": "M",
        "Asn": "N",
        "Pro": "P",
        "Gln": "Q",
        "Arg": "R",
        "Ser": "S",
        "Thr": "T",
        "Trp": "W",
        "Tyr": "Y",
        "Val": "V",
    }

    aa_change_fmt = aa_change.replace("p.", "")
    for l, s in aa_list.items():
        aa_change_fmt = aa_change_fmt.replace(l, s)

    return aa_change_fmt


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Creates an excel report containing sample and variants data"
    )

    parser.add_argument(
        "--sample-summary",
        "-s",
        required=True,
        dest="sample_summary",
        help="sample summary (.csv)",
    )
    parser.add_argument(
        "--variants-list",
        "-v",
        required=True,
        dest="variants_list",
        help="variant list (.csv)",
    )
    parser.add_argument(
        "--required-ref-coverage",
        "-r",
        default=95,
        type=int,
        dest="min_coverage",
        help="required reference coverage (as percentage) to call a valid assembly",
    )
    parser.add_argument("--output", "-o", required=True, help="Output file (.xlsx)")

    return parser.parse_args()


if __name__ == "__main__":
    main()
