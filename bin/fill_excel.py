#!/usr/bin/env python3
import sys
from openpyxl import load_workbook


def main(argv):
    row_start = 3
    wb = load_workbook(argv[-1])
    middle_index = int((len(argv) - 1) / 2)
    seq_names = argv[middle_index:-1]
    coverage = argv[0:middle_index]
    ws = wb.active

    for ix, seq_name in enumerate(seq_names):
        ws["C" + str(row_start)] = seq_name
        ws["D" + str(row_start)] = "betacoronavirus"
        ws["E" + str(row_start)] = "original"
        ws["I" + str(row_start)] = "Human"
        ws["S" + str(row_start)] = "Nanopore MinION"
        ws["T" + str(row_start)] = "ARTIC-nCoV-bioinformaticsSOP-v1.1.0"
        ws["U" + str(row_start)] = coverage[ix]
        ws[
            "V" + str(row_start)
        ] = "Molecular Medicine Laboratory, University of Magallanes"
        ws["W" + str(row_start)] = "Av. los Flamencos 01364"
        ws[
            "Y" + str(row_start)
        ] = "Centro Asistencial Docente y de Investigacion, Universidad de Magallanes"
        ws["Z" + str(row_start)] = "Av. los Flamencos 01364"
        ws[
            "AB" + str(row_start)
        ] = "Jorge González, Jacqueline Aldridge, Diego Alvarez, Ines Cid, Constanza Ceron, Roberto Uribe-Paredes, Marcelo Navarrete"
        row_start = row_start + 1

    wb.save("Submitted.xlsx")


if __name__ == "__main__":
    main(sys.argv[1:])
